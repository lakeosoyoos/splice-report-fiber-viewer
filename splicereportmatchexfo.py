#!/usr/bin/env python3
"""
splicereportmatchexfo.py — Splice QC report with EXFO-style bidirectional event matching
=========================================================================================

Extends the standard splice report with a second-pass B-direction scan to catch
events that only appear in the B-direction OTDR event table (no matching A event).
This matches EXFO's behavior of processing each direction independently and reporting
every event it finds, whether seen from one side or both.

PASS 1 (same as splice report):
  - For each fiber at each known splice closure position:
      Find A event → find matching B event → compute bidirectional loss
      Flag A-only events (no B match) if A loss >= threshold

PASS 2 (new):
  - For each fiber, scan ALL B-direction events above threshold
  - Convert each to A-frame coordinates
  - Skip any already caught in Pass 1
  - Match to nearest splice position (within 1.5 km)
  - If matching A event found: compute bidirectional loss → label A+B
  - If no A event: flag as B-only

CELL LABELS:
  325 .172        — standard A+B bidirectional splice (same as original report)
  325 .340 (B)    — B-direction only saw this event; A-direction had nothing
  325 .285 (A)    — A-direction only saw this event; no matching B entry

COLORS:
  Pink   — A+B bidirectional reburn (loss >= threshold)
  Red    — Break (1F reflective, clean cut)
  Orange — Broke (fiber terminates mid-span, crush/stress fracture)
  Blue   — B-fill (B-direction loss past a break, A-direction blind)
  Yellow — A-only (A saw it, B did not)
  Purple — B-only (B saw it, A did not)

USAGE
-----
    python splicereportmatchexfo.py A_DIR/ B_DIR/ --output report.xlsx

OPTIONS
    --output PATH    Output Excel file (default: splice_report_exfo.xlsx)
    --threshold dB   Flag threshold (default 0.150)
    --site-a NAME    A-end site name (default TUL)
    --site-b NAME    B-end site name (default BAR)
    --ribbon-size N  Fibers per ribbon (default 12)

REQUIREMENTS
    pip install numpy openpyxl
    sor_reader324802a.py must be in same directory.
"""

import os
import sys
import argparse
from collections import defaultdict

import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

from sor_reader324802a import parse_sor_full


# ═══════════════════════════════════════════════════════════════════════
#  DEFAULTS
# ═══════════════════════════════════════════════════════════════════════

REBURN_THRESHOLD = 0.150   # dB — flag anything at or above
NOMINAL_SPLICE   = 0.159   # dB expected per splice
RIBBON_SIZE      = 12      # fibers per ribbon
POSITION_TOL     = 1.5     # km tolerance for matching A↔B events
MIN_POP_SPLICE   = 20      # minimum fibers to define a splice position
END_REGION_KM    = 3.0     # last N km considered "end of fiber"


# ═══════════════════════════════════════════════════════════════════════
#  STEP 1 — Load all fibers
# ═══════════════════════════════════════════════════════════════════════

def load_all(dir_a, dir_b):
    fibers_a, fibers_b = {}, {}

    def extract_fiber_num(fn):
        base = fn.split('.')[0]
        base = base.split('_')[0]
        digits = ''.join(c for c in base if c.isdigit())
        return int(digits) if digits else None

    for fn in sorted(os.listdir(dir_a)):
        if not fn.lower().endswith('.sor'): continue
        r = parse_sor_full(os.path.join(dir_a, fn))
        if r:
            fnum = extract_fiber_num(fn)
            if fnum: fibers_a[fnum] = r

    if dir_b and os.path.isdir(dir_b):
        for fn in sorted(os.listdir(dir_b)):
            if not fn.lower().endswith('.sor'): continue
            r = parse_sor_full(os.path.join(dir_b, fn))
            if r:
                fnum = extract_fiber_num(fn)
                if fnum: fibers_b[fnum] = r

    return fibers_a, fibers_b


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2 — Discover splice closure positions from the A-direction population
# ═══════════════════════════════════════════════════════════════════════

def discover_splices(fibers_a):
    bins = defaultdict(list)
    for fnum, r in fibers_a.items():
        for e in r['events']:
            if e['dist_km'] < 1.0 or e['is_end']: continue
            if not e['type'].startswith('0F') and not e['type'].startswith('1F'): continue
            bk = round(e['dist_km'])
            bins[bk].append(e['dist_km'])

    splices = []
    for bk in sorted(bins.keys()):
        if len(bins[bk]) < MIN_POP_SPLICE: continue
        avg_pos = round(np.mean(bins[bk]), 2)
        splices.append({'bin': bk, 'position_km': avg_pos, 'count': len(bins[bk])})

    # Merge bins within 1 km of each other
    merged = []
    for sp in splices:
        if merged and abs(sp['position_km'] - merged[-1]['position_km']) < 1.0:
            if sp['count'] > merged[-1]['count']:
                merged[-1] = sp
        else:
            merged.append(sp)

    return merged


# ═══════════════════════════════════════════════════════════════════════
#  STEP 3 — Pass 1: Standard splice report analysis
#           (identical logic to splice_report_generator.py, plus A-only flagging)
# ═══════════════════════════════════════════════════════════════════════

def analyze_all(fibers_a, fibers_b, splices, threshold):
    """
    Pass 1: For each fiber at each known splice closure position:
      - Find A event → find matching B event → compute bidir loss → flag if above threshold
      - If no B match: flag A-only if A loss >= threshold (new vs original splice report)
      - Detect broke fibers and B-fill past breaks (same as original)

    event_source field:
      'bidir'  — both A and B direction saw it (standard splice)
      'a_only' — only A direction, no B match
      'broke'  — fiber terminates mid-span
      'bfill'  — B-direction fill past a break
    """
    results = {}

    # End-of-fiber distances for broke detection
    eof_a = {}
    for fnum, r in fibers_a.items():
        end = [e for e in r['events'] if e['is_end']]
        eof_a[fnum] = end[0]['dist_km'] if end else 999

    # Auto-detect span: top 25% median of all EOL distances
    eof_a_vals = sorted(eof_a.values())
    if eof_a_vals:
        top_quarter_a = eof_a_vals[int(len(eof_a_vals) * 0.75):]
        total_span_a = np.median(top_quarter_a)
    else:
        total_span_a = 0

    eof_b = {}
    for fnum, r in fibers_b.items():
        end = [e for e in r['events'] if e['is_end']]
        eof_b[fnum] = end[0]['dist_km'] if end else 999

    eof_b_vals = sorted([v for v in eof_b.values() if v < 999])
    if eof_b_vals:
        top_quarter_b = eof_b_vals[int(len(eof_b_vals) * 0.75):]
        total_span_b = np.median(top_quarter_b)
    else:
        total_span_b = 0

    for fnum, r in fibers_a.items():
        rb = fibers_b.get(fnum)
        b_span = None
        if rb:
            b_end = [e for e in rb['events'] if e['is_end']]
            b_span = b_end[0]['dist_km'] if b_end else total_span_b

        for si, sp in enumerate(splices):
            sp_km = sp['position_km']

            # ── Broke detection ──
            fiber_end = eof_a[fnum]
            a_plus_b = fiber_end + eof_b.get(fnum, 0) if fnum in eof_b else 0
            is_mid_span_break = (a_plus_b > 0 and
                                 abs(a_plus_b - total_span_a) < 3.0 and
                                 fiber_end < total_span_a - END_REGION_KM)

            if is_mid_span_break:
                # Mark as BROKE at the nearest splice to where it terminated
                nearest_splice = min(range(len(splices)),
                                     key=lambda i: abs(splices[i]['position_km'] - fiber_end))
                nearest_dist = abs(splices[nearest_splice]['position_km'] - fiber_end)
                if nearest_splice == si and nearest_dist < 2.0:
                    results[(fnum, si)] = {
                        'fiber': fnum, 'splice_idx': si,
                        'bidir_loss': None, 'a_loss': None, 'b_loss': None,
                        'bidir_dist': fiber_end,
                        'is_break': False, 'is_broke': True,
                        'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                        'is_flagged': True, 'event_source': 'broke',
                        'event_type': 'BROKE', 'label': f"{fnum} broke",
                    }
                # B-fill for splices past the break
                elif sp_km > fiber_end and rb and b_span:
                    b_evt = None
                    for e in rb['events']:
                        if e['dist_km'] < 1.0 or e['is_end']: continue
                        ef_from_a = b_span - e['dist_km']
                        if abs(ef_from_a - sp_km) < POSITION_TOL:
                            if b_evt is None or abs(ef_from_a - sp_km) < abs((b_span - b_evt['dist_km']) - sp_km):
                                b_evt = e
                    if b_evt is not None:
                        b_loss_val = abs(b_evt['splice_loss'])
                        if b_loss_val >= threshold:
                            loss_str = f"{b_loss_val:.3f}"
                            if loss_str.startswith('0.'): loss_str = loss_str[1:]
                            results[(fnum, si)] = {
                                'fiber': fnum, 'splice_idx': si,
                                'bidir_loss': b_loss_val, 'a_loss': None,
                                'b_loss': b_evt['splice_loss'],
                                'bidir_dist': b_span - b_evt['dist_km'],
                                'is_break': False, 'is_broke': False,
                                'is_bfill': True, 'is_a_only': False, 'is_b_only': False,
                                'is_flagged': True, 'event_source': 'bfill',
                                'event_type': b_evt['type'],
                                'label': f"{fnum} {loss_str} (B)",
                            }
                continue

            # ── Find A event near this splice ──
            ea = None
            for e in r['events']:
                if abs(e['dist_km'] - sp_km) < POSITION_TOL and e['dist_km'] > 1.0 and not e['is_end']:
                    if ea is None or abs(e['dist_km'] - sp_km) < abs(ea['dist_km'] - sp_km):
                        ea = e

            if ea is None:
                continue

            # ── Find matching B event ──
            eb = None
            b_loss = None
            b_from_a = None
            if rb and b_span:
                for e in rb['events']:
                    if e['dist_km'] < 1.0 or e['is_end']: continue
                    ef_from_a = b_span - e['dist_km']
                    if abs(ef_from_a - ea['dist_km']) < POSITION_TOL:
                        if eb is None or abs(ef_from_a - ea['dist_km']) < abs((b_span - eb['dist_km']) - ea['dist_km']):
                            eb = e
                            b_loss = e['splice_loss']
                            b_from_a = ef_from_a

            # ── A-only: no B match ──
            # B event table had no entry within tolerance — estimate bidir as A/2
            if b_loss is None:
                a_loss_abs = abs(ea['splice_loss'])
                if a_loss_abs / 2.0 >= threshold:
                    est_bidir = round(a_loss_abs / 2.0, 3)
                    loss_str = f"{a_loss_abs:.3f}"
                    if loss_str.startswith('0.'): loss_str = loss_str[1:]
                    est_str = f"{est_bidir:.3f}"
                    if est_str.startswith('0.'): est_str = est_str[1:]
                    # Mark whether estimated bidir still exceeds threshold
                    bidir_flag = '⚠' if est_bidir >= threshold else '~'
                    results[(fnum, si)] = {
                        'fiber': fnum, 'splice_idx': si,
                        'bidir_loss': None, 'a_loss': ea['splice_loss'], 'b_loss': None,
                        'bidir_dist': ea['dist_km'],
                        'est_bidir': est_bidir,
                        'est_bidir_flagged': est_bidir >= threshold,
                        'is_break': False, 'is_broke': False,
                        'is_bfill': False, 'is_a_only': True, 'is_b_only': False,
                        'is_flagged': True, 'event_source': 'a_only',
                        'event_type': ea['type'],
                        'label': f"{fnum} {loss_str}(A) {bidir_flag}{est_str}bd",
                    }
                continue

            # ── A+B bidirectional ──
            bidir_loss = round((ea['splice_loss'] + b_loss) / 2.0, 4)
            bidir_dist = round((ea['dist_km'] + b_from_a) / 2.0, 4)

            is_reflective = ea['type'].startswith('1F')
            has_weak_fresnel = ea['reflection'] < -30.0
            is_break = is_reflective and has_weak_fresnel and ea['dist_km'] < (total_span_a - END_REGION_KM)

            is_flagged = (abs(bidir_loss) >= threshold) or is_break
            if not is_flagged:
                continue

            if is_break:
                offset_m = round((bidir_dist - sp_km) * 1000, 1)
                label = f"{fnum} BREAK {bidir_loss:.3f} ({abs(offset_m):.0f}m from splice)"
            else:
                loss_str = f"{bidir_loss:.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                label = f"{fnum} {loss_str}"

            results[(fnum, si)] = {
                'fiber': fnum, 'splice_idx': si,
                'bidir_loss': bidir_loss,
                'a_loss': ea['splice_loss'], 'b_loss': b_loss,
                'bidir_dist': bidir_dist,
                'is_break': is_break, 'is_broke': False,
                'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                'is_flagged': True, 'event_source': 'bidir',
                'event_type': ea['type'],
                'label': label,
                'fresnel': ea['reflection'] if is_reflective else None,
            }

    return results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4 — Pass 2: Scan all B-direction events not caught in Pass 1
# ═══════════════════════════════════════════════════════════════════════

def scan_b_events(fibers_a, fibers_b, splices, threshold, existing_results, total_span_a):
    """
    Pass 2: For every B-direction event above threshold that was NOT already
    caught in Pass 1, find the nearest splice position (within 1.5 km) and report it.

    This is how EXFO finds events like fiber 325's 0.340 dB entry that only
    exists in the B-direction event table with no matching A-direction event.

    Returns a dict of (fnum, si) -> result — same structure as analyze_all().
    Does NOT overwrite any existing_results entries.
    """
    new_results = {}

    for fnum, rb in fibers_b.items():
        ra = fibers_a.get(fnum)

        # B-direction span (EOL)
        b_end_events = [e for e in rb['events'] if e['is_end']]
        if not b_end_events:
            continue
        b_span = b_end_events[0]['dist_km']

        # A-direction EOL (to know if this fiber is broken)
        ra_end_km = total_span_a
        if ra:
            a_end = [e for e in ra['events'] if e['is_end']]
            if a_end:
                ra_end_km = a_end[0]['dist_km']

        for e in rb['events']:
            if e['dist_km'] < 1.0 or e['is_end']:
                continue

            b_loss_signed = e['splice_loss']
            b_loss_abs = abs(b_loss_signed)
            if b_loss_abs / 2.0 < threshold:
                continue

            # Convert B-frame position to A-frame
            a_frame_km = b_span - e['dist_km']
            if a_frame_km < 0.5:
                continue  # launch artifact near B-end

            # Find nearest splice position within tolerance
            nearest_si = None
            nearest_dist = float('inf')
            for si, sp in enumerate(splices):
                d = abs(sp['position_km'] - a_frame_km)
                if d < nearest_dist:
                    nearest_dist = d
                    nearest_si = si

            if nearest_si is None or nearest_dist > POSITION_TOL:
                continue  # not near any known splice position

            # Already caught by Pass 1?
            if (fnum, nearest_si) in existing_results:
                continue

            # Already found a better match in this pass?
            if (fnum, nearest_si) in new_results:
                existing_a_frame = new_results[(fnum, nearest_si)]['bidir_dist']
                if nearest_dist >= abs(splices[nearest_si]['position_km'] - existing_a_frame):
                    continue

            # Look for A-direction event near the same A-frame position
            a_evt = None
            if ra:
                for ae in ra['events']:
                    if ae['dist_km'] < 1.0 or ae['is_end']: continue
                    if abs(ae['dist_km'] - a_frame_km) < POSITION_TOL:
                        if a_evt is None or abs(ae['dist_km'] - a_frame_km) < abs(a_evt['dist_km'] - a_frame_km):
                            a_evt = ae

            if a_evt is not None:
                # A event exists — compute bidirectional
                bidir = round((a_evt['splice_loss'] + b_loss_signed) / 2.0, 4)
                if abs(bidir) < threshold:
                    continue
                loss_str = f"{abs(bidir):.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                new_results[(fnum, nearest_si)] = {
                    'fiber': fnum, 'splice_idx': nearest_si,
                    'bidir_loss': bidir,
                    'a_loss': a_evt['splice_loss'], 'b_loss': b_loss_signed,
                    'bidir_dist': a_frame_km,
                    'is_break': False, 'is_broke': False,
                    'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                    'is_flagged': True, 'event_source': 'bidir',
                    'event_type': a_evt['type'],
                    'label': f"{fnum} {loss_str}",
                }
            else:
                # No A event — B-only
                # A event table had no entry within tolerance — estimate bidir as B/2
                est_bidir = round(b_loss_abs / 2.0, 3)
                loss_str = f"{b_loss_abs:.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                est_str = f"{est_bidir:.3f}"
                if est_str.startswith('0.'): est_str = est_str[1:]
                bidir_flag = '⚠' if est_bidir >= threshold else '~'
                new_results[(fnum, nearest_si)] = {
                    'fiber': fnum, 'splice_idx': nearest_si,
                    'bidir_loss': None,
                    'a_loss': None, 'b_loss': b_loss_signed,
                    'bidir_dist': a_frame_km,
                    'est_bidir': est_bidir,
                    'est_bidir_flagged': est_bidir >= threshold,
                    'is_break': False, 'is_broke': False,
                    'is_bfill': False, 'is_a_only': False, 'is_b_only': True,
                    'is_flagged': True, 'event_source': 'b_only',
                    'event_type': e['type'],
                    'label': f"{fnum} {loss_str}(B) {bidir_flag}{est_str}bd",
                }

    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 5 — Group into ribbons and build cell values
# ═══════════════════════════════════════════════════════════════════════

def build_ribbon_data(results, n_fibers, ribbon_size, n_splices):
    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    grid = {}

    for (fnum, si), res in results.items():
        ri = (fnum - 1) // ribbon_size
        key = (ri, si)
        if key not in grid:
            grid[key] = []
        grid[key].append(res)

    cells = {}
    for (ri, si), res_list in grid.items():
        res_list.sort(key=lambda x: x['fiber'])

        # Group fibers with same loss and same source type
        groups = []
        for res in res_list:
            merged = False
            for g in groups:
                if (res['bidir_loss'] is not None and g['loss'] is not None and
                        abs(res['bidir_loss'] - g['loss']) < 0.002 and
                        not res['is_break'] and not res['is_broke'] and
                        not g['is_break'] and not g['is_broke'] and
                        res.get('event_source') == g.get('event_source')):
                    g['fibers'].append(res['fiber'])
                    merged = True
                    break
            if not merged:
                groups.append({
                    'fibers': [res['fiber']],
                    'loss': res['bidir_loss'],
                    'is_break': res['is_break'],
                    'is_broke': res['is_broke'],
                    'is_bfill': res.get('is_bfill', False),
                    'is_a_only': res.get('is_a_only', False),
                    'is_b_only': res.get('is_b_only', False),
                    'event_source': res.get('event_source', 'bidir'),
                    'label': res['label'],
                    'res': res,
                })

        # Build cell text — label shows source for A-only and B-only
        parts = []
        for g in groups:
            if g['is_broke']:
                parts.append(f"{g['fibers'][0]} broke")
            elif g['is_break']:
                parts.append(g['label'])
            elif g['is_a_only']:
                fib_str = ','.join(str(f) for f in g['fibers'])
                raw_loss = g['res']['a_loss']
                loss_abs = abs(raw_loss) if raw_loss is not None else 0
                loss_str = f"{loss_abs:.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str} (A)")
            elif g['is_b_only']:
                fib_str = ','.join(str(f) for f in g['fibers'])
                raw_loss = g['res']['b_loss']
                loss_abs = abs(raw_loss) if raw_loss is not None else 0
                loss_str = f"{loss_abs:.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str} (B)")
            elif g.get('is_bfill'):
                fib_str = ','.join(str(f) for f in g['fibers'])
                loss = g['loss']
                loss_str = f"{loss:.3f}" if loss is not None else "?"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str} (B-fill)")
            else:
                fib_str = ','.join(str(f) for f in g['fibers'])
                loss = g['loss']
                loss_str = f"{loss:.3f}" if loss is not None else "?"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str}")

        cell_text = ' '.join(parts)
        is_break = any(g['is_break'] for g in groups)
        is_broke = any(g['is_broke'] for g in groups)
        is_bfill = any(g.get('is_bfill', False) for g in groups)

        # Has a standard bidir reburn in this cell?
        has_standard_reburn = any(
            not g['is_break'] and not g['is_broke'] and
            not g.get('is_bfill') and not g.get('is_a_only') and
            not g.get('is_b_only')
            for g in groups
        )
        # A-only / B-only only drive color if no higher-priority event present
        is_a_only = (any(g.get('is_a_only', False) for g in groups) and
                     not is_break and not is_broke and not is_bfill and not has_standard_reburn)
        is_b_only = (any(g.get('is_b_only', False) for g in groups) and
                     not is_break and not is_broke and not is_bfill and not has_standard_reburn)

        # If estimated bidir still clears threshold, use a stronger shade
        est_bidir_flagged = any(g['res'].get('est_bidir_flagged', False) for g in groups
                                if g.get('is_a_only') or g.get('is_b_only'))

        max_loss = max((g['loss'] for g in groups if g['loss'] is not None), default=0)

        cells[(ri, si)] = {
            'text': cell_text,
            'is_break': is_break,
            'is_broke': is_broke,
            'is_bfill': is_bfill,
            'is_a_only': is_a_only,
            'is_b_only': is_b_only,
            'est_bidir_flagged': est_bidir_flagged,
            'max_loss': max_loss,
        }

    return cells


# ═══════════════════════════════════════════════════════════════════════
#  STEP 6 — Generate Excel
# ═══════════════════════════════════════════════════════════════════════

def ribbon_label(ri, ribbon_size, n_fibers):
    first = ri * ribbon_size + 1
    last = min(first + ribbon_size - 1, n_fibers)
    ribbon_num = ri + 1
    tube = ''
    if ri < 48:
        tube_letter = chr(ord('A') + ri // 2)
        tube_num = (ri % 2) + 1
        tube = f" ({tube_letter}{tube_num})"
    return f"Fiber {first}-{last} ({ribbon_num}){tube}"


def write_xlsx(cells, splices, n_fibers, ribbon_size, output_path, site_a, site_b, span_km):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Splice Report"

    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    n_splices = len(splices)

    # ── Styles ──
    hdr_font    = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font   = Font(size=8)
    ribbon_font = Font(size=9)
    a_km_font   = Font(bold=True, size=9, color="1F4E79")
    b_km_font   = Font(bold=True, size=9, color="8B0000")

    # Cell fill/font for each event type
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # A+B reburn
    break_fill  = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")   # break
    break_font  = Font(bold=True, size=8, color="FFFFFF")
    broke_fill  = PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid")   # broke
    broke_font  = Font(bold=True, size=8, color="FFFFFF")
    bfill_fill  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")   # B-fill past break
    bfill_font  = Font(size=8, color="1F4E79")
    aonly_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # A-only (yellow, est bidir OK)
    aonly_font  = Font(size=8, color="7F6000")
    aonly_fill2 = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # A-only (gold, est bidir >= threshold)
    aonly_font2 = Font(bold=True, size=8, color="4B3000")
    bonly_fill  = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")   # B-only (lavender, est bidir OK)
    bonly_font  = Font(size=8, color="4B0082")
    bonly_fill2 = PatternFill(start_color="C084FC", end_color="C084FC", fill_type="solid")   # B-only (purple, est bidir >= threshold)
    bonly_font2 = Font(bold=True, size=8, color="1A0033")

    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # ── Row 1: A→B distances (km / ft) ──
    ws.cell(row=1, column=2, value="A→B:").font = a_km_font
    ws.cell(row=2, column=2, value="B→A:").font = b_km_font
    for si, sp in enumerate(splices):
        col = si + 3
        km = sp['position_km']
        ft = km * 3280.84
        b_km = span_km - km
        b_ft = b_km * 3280.84
        c1 = ws.cell(row=1, column=col, value=f"{km:.2f}km / {ft:,.0f}ft")
        c1.font = a_km_font
        c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=2, column=col, value=f"{b_km:.2f}km / {b_ft:,.0f}ft")
        c2.font = b_km_font
        c2.alignment = Alignment(horizontal='center')
    end_col = n_splices + 3
    ws.cell(row=1, column=end_col, value=f"{span_km:.2f}km / {span_km*3280.84:,.0f}ft").font = a_km_font
    ws.cell(row=2, column=end_col, value="0.00km / 0ft").font = b_km_font

    # ── Row 3: Headers ──
    ws.cell(row=3, column=1, value="Ribbon").font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill
    ws.cell(row=3, column=2, value=f"ILA:{site_a}").font = hdr_font
    ws.cell(row=3, column=2).fill = hdr_fill
    for si in range(n_splices):
        col = si + 3
        cell = ws.cell(row=3, column=col, value=f"Splice {si+1}")
        cell.font = hdr_font
        cell.fill = hdr_fill
    ws.cell(row=3, column=end_col, value=f"ILA:{site_b}").font = hdr_font
    ws.cell(row=3, column=end_col).fill = hdr_fill

    # ── Data rows ──
    for ri in range(n_ribbons):
        row = ri + 4
        ws.cell(row=row, column=1, value=ribbon_label(ri, ribbon_size, n_fibers)).font = ribbon_font
        for si in range(n_splices):
            col = si + 3
            key = (ri, si)
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')

            if key in cells:
                cd = cells[key]
                cell.value = cd['text']
                if cd['is_break']:
                    cell.fill = break_fill
                    cell.font = break_font
                elif cd['is_broke']:
                    cell.fill = broke_fill
                    cell.font = broke_font
                elif cd.get('is_bfill'):
                    cell.fill = bfill_fill
                    cell.font = bfill_font
                elif cd.get('is_b_only'):
                    if cd.get('est_bidir_flagged'):
                        cell.fill = bonly_fill2
                        cell.font = bonly_font2
                    else:
                        cell.fill = bonly_fill
                        cell.font = bonly_font
                elif cd.get('is_a_only'):
                    if cd.get('est_bidir_flagged'):
                        cell.fill = aonly_fill2
                        cell.font = aonly_font2
                    else:
                        cell.fill = aonly_fill
                        cell.font = aonly_font
                else:
                    cell.fill = red_fill
                    cell.font = data_font

    # ── Legend sheet ──
    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions['A'].width = 14
    ws_leg.column_dimensions['B'].width = 65
    legend_items = [
        ("Pink",       "FFC7CE", "000000", "A+B — Bidirectional reburn: both directions confirmed, bidir loss >= threshold. Needs re-splice."),
        ("Red",        "FF4444", "FFFFFF", "Break — 1F reflective event (clean cut, glass-to-air Fresnel reflection). label: 'BREAK'"),
        ("Orange",     "FF8800", "FFFFFF", "Broke — fiber trace terminates mid-span, no reflection (crush / stress fracture). label: 'broke'"),
        ("Blue",       "BDD7EE", "1F4E79", "B-fill — B-direction loss used past a break where A-direction is blind. label: '(B-fill)'"),
        ("Lt. Yellow", "FFF2CC", "7F6000", "A-only, est bidir OK — A saw it, no B entry. Estimated bidir (A/2) is below threshold. label: 'F# .xxx(A) ~.xxxbd'"),
        ("Gold",       "FFD700", "4B3000", "A-only, est bidir HIGH — A saw it, no B entry. Estimated bidir (A/2) still exceeds threshold. label: 'F# .xxx(A) ⚠.xxxbd'"),
        ("Lavender",   "E8D5F5", "4B0082", "B-only, est bidir OK — B saw it, no A entry. Estimated bidir (B/2) is below threshold. label: 'F# .xxx(B) ~.xxxbd'"),
        ("Purple",     "C084FC", "1A0033", "B-only, est bidir HIGH — B saw it, no A entry. Estimated bidir (B/2) still exceeds threshold. label: 'F# .xxx(B) ⚠.xxxbd'"),
    ]
    ws_leg.cell(row=1, column=1, value="Color").font = Font(bold=True, size=10)
    ws_leg.cell(row=1, column=2, value="Meaning").font = Font(bold=True, size=10)
    for i, (name, fc, tc, desc) in enumerate(legend_items, 2):
        c = ws_leg.cell(row=i, column=1, value=name)
        c.fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
        c.font = Font(bold=True, size=9, color=tc)
        ws_leg.cell(row=i, column=2, value=desc).font = Font(size=9)

    # ── Column widths ──
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    for si in range(n_splices + 1):
        col_letter = openpyxl.utils.get_column_letter(si + 3)
        ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = 'C4'

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description='Splice QC report with EXFO-style bidirectional event matching.',
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('dir_a', help='A-direction SOR files directory')
    ap.add_argument('dir_b', nargs='?', help='B-direction SOR files directory')
    ap.add_argument('--output', '-o', default='splice_report_exfo.xlsx')
    ap.add_argument('--threshold', type=float, default=REBURN_THRESHOLD,
                    help=f'Flag threshold in dB (default {REBURN_THRESHOLD})')
    ap.add_argument('--ribbon-size', type=int, default=RIBBON_SIZE)
    ap.add_argument('--site-a', default='TUL')
    ap.add_argument('--site-b', default='BAR')
    ap.add_argument('--span-km', type=float, default=0,
                    help='Span distance in km (0 = auto-detect)')
    args = ap.parse_args()

    print("Loading SOR files...")
    fibers_a, fibers_b = load_all(args.dir_a, args.dir_b)
    n_fibers = max(fibers_a.keys()) if fibers_a else 0
    print(f"  A: {len(fibers_a)} fibers   B: {len(fibers_b)} fibers   max fiber #{n_fibers}")

    print("Discovering splice closure positions...")
    splices = discover_splices(fibers_a)
    print(f"  Found {len(splices)} splice closures:")
    for i, sp in enumerate(splices, 1):
        print(f"    Splice {i:2d}: {sp['position_km']:8.2f} km  ({sp['count']} fibers)")

    # Auto-detect span
    span_km = args.span_km
    if span_km == 0:
        all_ends = sorted([e['dist_km'] for r in fibers_a.values()
                           for e in r['events'] if e['is_end']])
        if all_ends:
            top_quarter = all_ends[int(len(all_ends) * 0.75):]
            span_km = round(np.median(top_quarter), 2)
    print(f"  Span: {span_km} km ({span_km * 3280.84:,.0f} ft)")

    print(f"\nPass 1: Analyzing {len(fibers_a)} fibers at {len(splices)} splice positions "
          f"(threshold={args.threshold:.3f} dB)...")
    results = analyze_all(fibers_a, fibers_b, splices, args.threshold)
    n_p1_bidir  = sum(1 for r in results.values() if r.get('event_source') == 'bidir')
    n_p1_aonly  = sum(1 for r in results.values() if r.get('is_a_only'))
    n_p1_broke  = sum(1 for r in results.values() if r['is_broke'])
    n_p1_break  = sum(1 for r in results.values() if r['is_break'])
    n_p1_bfill  = sum(1 for r in results.values() if r.get('is_bfill'))
    print(f"  Pass 1 results: {len(results)} events")
    print(f"    A+B bidir:  {n_p1_bidir}")
    print(f"    A-only:     {n_p1_aonly}")
    print(f"    Breaks:     {n_p1_break}")
    print(f"    Broke:      {n_p1_broke}")
    print(f"    B-fill:     {n_p1_bfill}")

    print(f"\nPass 2: Scanning B-direction events not caught in Pass 1...")
    b_results = scan_b_events(fibers_a, fibers_b, splices, args.threshold, results, span_km)
    n_p2_bidir = sum(1 for r in b_results.values() if r.get('event_source') == 'bidir')
    n_p2_bonly = sum(1 for r in b_results.values() if r.get('is_b_only'))
    print(f"  Pass 2 results: {len(b_results)} additional events")
    print(f"    A+B (via B-scan): {n_p2_bidir}")
    print(f"    B-only:           {n_p2_bonly}")

    # Merge — Pass 1 takes priority
    all_results = {**results, **b_results}

    n_total   = len(all_results)
    n_bidir   = sum(1 for r in all_results.values() if r.get('event_source') == 'bidir')
    n_a_only  = sum(1 for r in all_results.values() if r.get('is_a_only'))
    n_b_only  = sum(1 for r in all_results.values() if r.get('is_b_only'))
    n_breaks  = sum(1 for r in all_results.values() if r['is_break'])
    n_broke   = sum(1 for r in all_results.values() if r['is_broke'])
    n_bfill   = sum(1 for r in all_results.values() if r.get('is_bfill'))
    n_reburn  = n_bidir - n_breaks

    print(f"\nBuilding ribbon grid...")
    cells = build_ribbon_data(all_results, n_fibers, args.ribbon_size, len(splices))
    print(f"  {len(cells)} cells with flagged events")

    print(f"Writing Excel report...")
    write_xlsx(cells, splices, n_fibers, args.ribbon_size, args.output,
               args.site_a, args.site_b, span_km)

    print(f"\n{'═'*60}")
    print(f"  SPLICE REPORT (EXFO-MATCH) COMPLETE")
    print(f"{'═'*60}")
    print(f"  Fibers:       {n_fibers}")
    print(f"  Splices:      {len(splices)}")
    print(f"  Span:         {span_km} km")
    print(f"  Threshold:    {args.threshold:.3f} dB")
    print(f"  ──────────────────────────────────")
    print(f"  A+B reburns:  {n_reburn}  (pink)   — both directions, bidir >= threshold")
    print(f"  Breaks:       {n_breaks}  (red)    — 1F reflective event")
    print(f"  Broke:        {n_broke}  (orange) — trace terminates mid-span")
    print(f"  B-fill:       {n_bfill}  (blue)   — B-direction past a break")
    print(f"  A-only:       {n_a_only}  (yellow) — A saw it, B did not")
    print(f"  B-only:       {n_b_only}  (purple) — B saw it, A did not  ← EXFO extra")
    print(f"  ──────────────────────────────────")
    print(f"  Total:        {n_total}")
    print(f"  Output:       {args.output}")
    print()


if __name__ == '__main__':
    main()
